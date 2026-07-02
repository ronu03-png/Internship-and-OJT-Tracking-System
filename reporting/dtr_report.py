"""
dtr_report.py
=============
Automated DTR (Daily Time Record) Excel report generator for the
Internship & OJT Tracking System (CHED CMO No. 104, s. 2017).

Pulls APPROVED DTR rows for a given internship straight from the PostgreSQL /
Supabase database using psycopg2, then compiles a presentation-ready .xlsx via
openpyxl featuring:
  * Institution header block
  * Student + HTE metadata
  * Bordered table with bold headers
  * Automatic total-hours calculation
  * Physical/digital signature slots for the Student and HTE Supervisor

Usage
-----
    python dtr_report.py --internship <UUID> [--out report.xlsx]

Environment (see .env.example)
------------------------------
    DATABASE_URL   postgresql://user:pass@host:5432/postgres
    (or discrete: PGHOST, PGPORT, PGDATABASE, PGUSER, PGPASSWORD)

    SCHOOL_NAME      optional header line
    ACADEMIC_YEAR    optional header line
"""

from __future__ import annotations

import argparse
import os
import sys
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from typing import Optional

import psycopg2
import psycopg2.extras

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# --------------------------------------------------------------------------- #
# Styling constants
# --------------------------------------------------------------------------- #
THIN = Side(style="thin", color="9AA5B1")
MEDIUM = Side(style="medium", color="334155")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BOX_MED = Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)

HEADER_FILL = PatternFill("solid", fgColor="1D4ED8")   # brand blue
SUBTLE_FILL = PatternFill("solid", fgColor="EFF4FF")
TOTAL_FILL = PatternFill("solid", fgColor="DCFCE7")    # light green

WHITE_BOLD = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=16, color="0F172A")
SUB_FONT = Font(size=10, color="475569")
LABEL_FONT = Font(bold=True, size=10, color="0F172A")
CELL_FONT = Font(size=10, color="1E293B")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

COLUMNS = [
    ("No.", 6),
    ("Date", 14),
    ("Day", 12),
    ("Time In", 12),
    ("Time Out", 12),
    ("Rendered Hours", 16),
    ("Remarks", 26),
]


# --------------------------------------------------------------------------- #
# Data structures
# --------------------------------------------------------------------------- #
@dataclass
class Meta:
    student_name: str
    course: Optional[str]
    hte_name: str
    supervisor_name: Optional[str]
    required_hours: Decimal
    hours_rendered: Decimal
    start_date: Optional[date]
    end_date: Optional[date]


@dataclass
class DtrRow:
    log_date: date
    time_in: Optional[datetime]
    time_out: Optional[datetime]
    rendered_hours: Decimal
    remarks: Optional[str]


# --------------------------------------------------------------------------- #
# Database access
# --------------------------------------------------------------------------- #
def get_connection():
    """Open a psycopg2 connection from DATABASE_URL or discrete PG* vars."""
    dsn = os.getenv("DATABASE_URL")
    if dsn:
        return psycopg2.connect(dsn)
    return psycopg2.connect(
        host=os.getenv("PGHOST", "localhost"),
        port=os.getenv("PGPORT", "5432"),
        dbname=os.getenv("PGDATABASE", "postgres"),
        user=os.getenv("PGUSER", "postgres"),
        password=os.getenv("PGPASSWORD", ""),
    )


def fetch_meta(conn, internship_id: str) -> Meta:
    sql = """
        select
            stu.full_name           as student_name,
            i.course                as course,
            h.name                  as hte_name,
            sup.full_name           as supervisor_name,
            i.required_hours        as required_hours,
            i.hours_rendered        as hours_rendered,
            i.start_date            as start_date,
            i.end_date              as end_date
        from internships i
        join profiles stu on stu.id = i.student_id
        join htes h       on h.id = i.hte_id
        left join profiles sup on sup.id = i.supervisor_id
        where i.id = %s
    """
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute(sql, (internship_id,))
        row = cur.fetchone()
        if not row:
            raise SystemExit(f"No internship found for id {internship_id!r}")
        return Meta(**row)


def fetch_dtr_rows(conn, internship_id: str) -> list[DtrRow]:
    """Only APPROVED rows are audit-eligible."""
    sql = """
        select log_date, time_in, time_out, rendered_hours, remarks
        from dtr_logs
        where internship_id = %s and status = 'approved'
        order by log_date asc
    """
    with conn.cursor() as cur:
        cur.execute(sql, (internship_id,))
        return [
            DtrRow(
                log_date=r[0],
                time_in=r[1],
                time_out=r[2],
                rendered_hours=r[3] or Decimal(0),
                remarks=r[4],
            )
            for r in cur.fetchall()
        ]


# --------------------------------------------------------------------------- #
# Worksheet builders
# --------------------------------------------------------------------------- #
def _fmt_time(ts: Optional[datetime]) -> str:
    return ts.strftime("%I:%M %p") if ts else "-"


def build_header(ws: Worksheet, meta: Meta, last_col: int) -> int:
    """Institution + student/HTE metadata block. Returns next free row index."""
    span = get_column_letter(last_col)

    ws.merge_cells(f"A1:{span}1")
    ws["A1"] = os.getenv("SCHOOL_NAME", "University of Example Philippines")
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER

    ws.merge_cells(f"A2:{span}2")
    ws["A2"] = "Daily Time Record (DTR) — On-the-Job Training"
    ws["A2"].font = Font(bold=True, size=12, color="1D4ED8")
    ws["A2"].alignment = CENTER

    ws.merge_cells(f"A3:{span}3")
    ws["A3"] = (
        f"Academic Year {os.getenv('ACADEMIC_YEAR', '2025-2026')} · "
        "In compliance with CHED CMO No. 104, s. 2017"
    )
    ws["A3"].font = SUB_FONT
    ws["A3"].alignment = CENTER

    row = 5
    info = [
        ("Student Name:", meta.student_name, "Host Establishment:", meta.hte_name),
        ("Course:", meta.course or "-", "HTE Supervisor:", meta.supervisor_name or "-"),
        (
            "Coverage:",
            f"{meta.start_date or '-'} to {meta.end_date or '-'}",
            "Required Hours:",
            f"{meta.required_hours:g}",
        ),
    ]
    for r_off, (l1, v1, l2, v2) in enumerate(info):
        r = row + r_off
        ws.cell(r, 1, l1).font = LABEL_FONT
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        ws.cell(r, 2, v1).font = CELL_FONT
        ws.cell(r, 4, l2).font = LABEL_FONT
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=last_col)
        ws.cell(r, 5, v2).font = CELL_FONT

    return row + len(info) + 1


def build_table(ws: Worksheet, start_row: int, rows: list[DtrRow]) -> tuple[int, Decimal]:
    """Render the bordered DTR table. Returns (total_row_index, total_hours)."""
    # Header row
    for c_idx, (title, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(start_row, c_idx, title)
        cell.font = WHITE_BOLD
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BOX

    total = Decimal(0)
    r = start_row + 1
    for i, row in enumerate(rows, start=1):
        total += row.rendered_hours or Decimal(0)
        values = [
            i,
            row.log_date.strftime("%Y-%m-%d"),
            row.log_date.strftime("%A"),
            _fmt_time(row.time_in),
            _fmt_time(row.time_out),
            float(row.rendered_hours or 0),
            row.remarks or "",
        ]
        for c_idx, val in enumerate(values, start=1):
            cell = ws.cell(r, c_idx, val)
            cell.font = CELL_FONT
            cell.border = BOX
            if c_idx in (1, 2, 3, 4, 5, 6):
                cell.alignment = CENTER
            else:
                cell.alignment = LEFT
            if i % 2 == 0:
                cell.fill = SUBTLE_FILL
        if c_idx:  # numeric hours format
            ws.cell(r, 6).number_format = "0.00"
        r += 1

    if not rows:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(COLUMNS))
        empty = ws.cell(r, 1, "No approved DTR records found.")
        empty.alignment = CENTER
        empty.font = Font(italic=True, color="94A3B8")
        empty.border = BOX
        r += 1

    # Total row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    label = ws.cell(r, 1, "TOTAL APPROVED HOURS")
    label.font = Font(bold=True, size=11, color="065F46")
    label.alignment = RIGHT
    label.fill = TOTAL_FILL
    label.border = BOX_MED

    total_cell = ws.cell(r, 6, float(total))
    total_cell.font = Font(bold=True, size=11, color="065F46")
    total_cell.alignment = CENTER
    total_cell.number_format = "0.00"
    total_cell.fill = TOTAL_FILL
    total_cell.border = BOX_MED

    ws.cell(r, 7, "").border = BOX_MED
    return r, total


def build_signatures(ws: Worksheet, start_row: int, meta: Meta) -> None:
    """Physical/digital signature slots for Student and HTE Supervisor."""
    r = start_row + 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(COLUMNS))
    note = ws.cell(
        r, 1,
        "I hereby certify that the above record is true and correct.",
    )
    note.font = Font(italic=True, size=9, color="475569")
    note.alignment = LEFT

    sig_row = r + 3
    # Student block (cols 1-3), Supervisor block (cols 5-7)
    for col_start, name, role in (
        (1, meta.student_name, "Student / Trainee"),
        (5, meta.supervisor_name or "________________________", "HTE Supervisor"),
    ):
        col_end = col_start + 2
        line = ws.cell(sig_row, col_start, name)
        ws.merge_cells(
            start_row=sig_row, start_column=col_start,
            end_row=sig_row, end_column=col_end,
        )
        line.alignment = CENTER
        line.font = Font(bold=True, size=10, color="0F172A")
        line.border = Border(bottom=MEDIUM)

        cap = ws.cell(sig_row + 1, col_start, f"{role} (Signature over Printed Name / Date)")
        ws.merge_cells(
            start_row=sig_row + 1, start_column=col_start,
            end_row=sig_row + 1, end_column=col_end,
        )
        cap.alignment = CENTER
        cap.font = Font(size=8, color="64748B")


def style_sheet(ws: Worksheet) -> None:
    ws.sheet_view.showGridLines = False
    for c_idx, (_, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = width
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.print_title_rows = "1:3"


# --------------------------------------------------------------------------- #
# Orchestration
# --------------------------------------------------------------------------- #
def generate_report(internship_id: str, out_path: str) -> str:
    conn = get_connection()
    try:
        meta = fetch_meta(conn, internship_id)
        rows = fetch_dtr_rows(conn, internship_id)
    finally:
        conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "DTR"
    style_sheet(ws)

    last_col = len(COLUMNS)
    table_start = build_header(ws, meta, last_col)
    total_row, _total = build_table(ws, table_start, rows)
    build_signatures(ws, total_row, meta)

    wb.save(out_path)
    return out_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate an OJT DTR Excel report.")
    parser.add_argument("--internship", required=True, help="internship UUID")
    parser.add_argument("--out", default=None, help="output .xlsx path")
    args = parser.parse_args()

    out = args.out or f"DTR_{args.internship[:8]}.xlsx"
    path = generate_report(args.internship, out)
    print(f"Report written to: {path}", file=sys.stderr)


if __name__ == "__main__":
    main()
